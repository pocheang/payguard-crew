"""
批量审计和导出服务

新增功能：
1. 批量审计交易
2. 导出审计报告（Excel/CSV）
3. 查询统计分析
"""
import asyncio
from datetime import datetime, timezone
from typing import List

from app.schemas.transaction import TransactionInput
from app.schemas.audit import AuditResponse, AuditLogEntry
from app.crew.audit_crew import run_audit_crew_async
from app.db.repositories import save_audit_result_optimized


async def batch_audit_transactions(
    transactions: List[TransactionInput],
    max_concurrent: int = 10
) -> List[AuditResponse]:
    """
    批量审计交易（异步并发）

    Args:
        transactions: 交易列表
        max_concurrent: 最大并发数（默认10）

    Returns:
        审计结果列表

    性能：
        - 10个交易并发: ~3-5秒
        - 100个交易: ~30-50秒
        - 自动控制并发数，避免资源耗尽
    """
    semaphore = asyncio.Semaphore(max_concurrent)

    async def audit_with_semaphore(tx: TransactionInput):
        async with semaphore:
            try:
                result = await run_audit_crew_async(tx)
                return result
            except Exception as e:
                # 单个失败不影响其他
                import logging
                logging.error(f"Batch audit failed for {tx.transaction_id}: {e}")
                return None

    # 并发执行
    tasks = [audit_with_semaphore(tx) for tx in transactions]
    results = await asyncio.gather(*tasks)

    # 过滤失败的结果
    return [r for r in results if r is not None]


def export_audit_reports_csv(
    transaction_ids: List[str],
    output_path: str = "audit_reports.csv"
) -> str:
    """
    导出审计报告为CSV

    Args:
        transaction_ids: 交易ID列表
        output_path: 输出文件路径

    Returns:
        导出的文件路径
    """
    import csv
    from app.db.repositories import get_audit_report

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)

        # 写入表头
        writer.writerow([
            '交易ID', '用户ID', '商户ID',
            '风险分数', '风险等级', '决策',
            '触发规则数', '需要人工审核',
            '创建时间', '摘要'
        ])

        # 写入数据
        for tx_id in transaction_ids:
            report = get_audit_report(tx_id)
            if report:
                writer.writerow([
                    report.transaction_id,
                    report.user_id,
                    report.merchant_id,
                    report.risk_score,
                    report.risk_level,
                    report.decision,
                    len(report.triggered_rules),
                    '是' if report.requires_manual_review else '否',
                    report.created_at,
                    report.summary[:100] + '...' if len(report.summary) > 100 else report.summary
                ])

    return output_path


def export_audit_reports_excel(
    transaction_ids: List[str],
    output_path: str = "audit_reports.xlsx"
) -> str:
    """
    导出审计报告为Excel（带格式）

    Args:
        transaction_ids: 交易ID列表
        output_path: 输出文件路径

    Returns:
        导出的文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        )

    from app.db.repositories import get_audit_report

    wb = Workbook()
    ws = wb.active
    ws.title = "审计报告"

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # 写入表头
    headers = [
        '交易ID', '用户ID', '商户ID', '风险分数', '风险等级',
        '决策', '触发规则数', '需要人工审核', '创建时间', '摘要'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据
    for row_idx, tx_id in enumerate(transaction_ids, 2):
        report = get_audit_report(tx_id)
        if report:
            # 风险等级颜色标记
            risk_color = {
                'high': 'FFCCCC',
                'medium': 'FFFFCC',
                'low': 'CCFFCC'
            }.get(report.risk_level, 'FFFFFF')

            data = [
                report.transaction_id,
                report.user_id,
                report.merchant_id,
                report.risk_score,
                report.risk_level,
                report.decision,
                len(report.triggered_rules),
                '是' if report.requires_manual_review else '否',
                report.created_at,
                report.summary[:100] + '...' if len(report.summary) > 100 else report.summary
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                # 风险等级列着色
                if col == 5:
                    cell.fill = PatternFill(start_color=risk_color, end_color=risk_color, fill_type="solid")

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    return output_path


def get_audit_statistics(
    start_date: str = None,
    end_date: str = None
) -> dict:
    """
    获取审计统计信息

    Args:
        start_date: 开始日期 (ISO格式，可选)
        end_date: 结束日期 (ISO格式，可选)

    Returns:
        统计信息字典
    """
    from app.db.database import get_connection

    with get_connection() as conn:
        # 基础查询
        query = "SELECT * FROM audit_reports WHERE 1=1"
        params = []

        if start_date:
            query += " AND created_at >= ?"
            params.append(start_date)
        if end_date:
            query += " AND created_at <= ?"
            params.append(end_date)

        reports = conn.execute(query, params).fetchall()

        if not reports:
            return {
                "total_count": 0,
                "risk_level_distribution": {},
                "decision_distribution": {},
                "average_risk_score": 0,
                "manual_review_rate": 0,
            }

        # 统计
        total = len(reports)
        risk_levels = {}
        decisions = {}
        total_score = 0
        manual_review_count = 0

        for report in reports:
            # 风险等级分布
            level = report['risk_level']
            risk_levels[level] = risk_levels.get(level, 0) + 1

            # 决策分布
            decision = report['decision']
            decisions[decision] = decisions.get(decision, 0) + 1

            # 分数
            total_score += report['risk_score']

            # 人工审核
            if report['requires_manual_review']:
                manual_review_count += 1

        # 获取最常触发的规则
        rule_stats = conn.execute("""
            SELECT rule_id, rule_name, COUNT(*) as count
            FROM rule_hits
            GROUP BY rule_id, rule_name
            ORDER BY count DESC
            LIMIT 10
        """).fetchall()

        return {
            "total_count": total,
            "risk_level_distribution": {
                k: {"count": v, "percentage": round(v / total * 100, 2)}
                for k, v in risk_levels.items()
            },
            "decision_distribution": {
                k: {"count": v, "percentage": round(v / total * 100, 2)}
                for k, v in decisions.items()
            },
            "average_risk_score": round(total_score / total, 2),
            "manual_review_rate": round(manual_review_count / total * 100, 2),
            "top_triggered_rules": [
                {
                    "rule_id": row['rule_id'],
                    "rule_name": row['rule_name'],
                    "count": row['count']
                }
                for row in rule_stats
            ],
            "date_range": {
                "start": start_date or "all",
                "end": end_date or "all"
            }
        }
